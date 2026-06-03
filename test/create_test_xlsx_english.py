#!/usr/bin/env python3
"""
Script to create a test XLSX file with English headers for salary import
"""
import os

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# English headers. One row = one payment, grouped by the "Salary" notation.
headers = [
    'Salary',
    'Payment ref',
    'First name',
    'Last name',
    'Payment date',
    'Payment type',
    'Bank account',
    'Amount paid',
    'Amount CHF',
    'Total salary CHF',
    'Label',
    'Start date',
    'End date',
    'Paid'
]

# Write headers
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Test data - mono-payment salaries on a CHF account (paid == CHF == total)
data = [
    ['2026-01-1', '2026-01-1-CHF', 'John', 'Doe', '2026-01-15', 'VIR', 'POSTE_CH', 4500.00, 4500.00, 4500.00, 'Salary January 2026', '2026-01-01', '2026-01-31', 'yes'],
    ['2026-01-2', '2026-01-2-CHF', 'Jane', 'Smith', '2026-01-15', 'VIR', 'POSTE_CH', 4200.00, 4200.00, 4200.00, 'Salary January 2026', '2026-01-01', '2026-01-31', 'yes'],
]

# Write data
for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Save
output_path = os.path.join(os.path.dirname(__file__), 'salaires_test_english.xlsx')
wb.save(output_path)
print(f"Created: {output_path}")
